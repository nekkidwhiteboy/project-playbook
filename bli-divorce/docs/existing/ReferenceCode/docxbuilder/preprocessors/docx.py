from pathlib import Path
from typing import Any, Dict

from django.core.files import File
from openpyxl import load_workbook

from . import BasePreprocessor


class CSOWPreprocessor(BasePreprocessor):
    data_path = Path(__file__).resolve().parent / "csow_data"

    def __call__(self, context: Dict[str, Any], template: File) -> Dict[str, Any]:
        context["get_csow_data"] = lambda: self._get_data(context)

        return context

    def _get_data(self, context):
        pet = self._get_basic_info("Pet", context)
        resp = self._get_basic_info("Resp", context)

        data = {"ChildInfo": context["ChildInfo"], "TotalWAI": pet["WAI"] + resp["WAI"]}

        pet["PercentageWAI"] = pet["WAI"] / data["TotalWAI"]
        resp["PercentageWAI"] = resp["WAI"] / data["TotalWAI"]

        data["Obligation"] = self._calc_obligation(
            data["TotalWAI"], context["FilingState"], len(context["ChildInfo"])
        )
        data["TotalSupport"] = (
            _to_weekly(context["ChildInsuranceCost"] or 0.0)
            + data["Obligation"]
            + pet["ChildCare"]
            + resp["ChildCare"]
        )

        if context["RespSex"] == "Male":
            resp_id = "Father"
        elif context["RespSex"] == "Female":
            resp_id = "Mother"
        else:
            resp_id = "Respondent"

        if context["PetSex"] == "Male":
            pet_id = "Father"
        elif context["PetSex"] == "Female":
            pet_id = "Mother"
        else:
            pet_id = "Respondent"

        # If one party pays controlled expenses (due to equal parenting time)
        #   the other party pays support, unless their credit exceeds the amount due
        if pet["IsControlledExpensePayer"]:
            pet["RecSupport"] = 0
            resp["PTC"] = self._get_ptc(resp, data)
            resp["RecSupport"] = self._get_rec_support(resp, data["TotalSupport"])
            # Credit exceeds amount due
            if resp["RecSupport"] < 0.0:
                pet["RecSupport"] = -resp["RecSupport"]
                resp["RecSupport"] = 0.0
                data["Message"] = (
                    f"Because the {resp_id}'s credits exceed their support obligation, "
                    f"the court should order the {pet_id} to pay the amount on line 8."
                )
        elif resp["IsControlledExpensePayer"]:
            resp["RecSupport"] = 0
            pet["PTC"] = self._get_ptc(pet, data)
            pet["RecSupport"] = self._get_rec_support(pet, data["TotalSupport"])
            # Credit exceeds amount due
            if pet["RecSupport"] < 0.0:
                resp["RecSupport"] = -pet["RecSupport"]
                pet["RecSupport"] = 0.0
                data["Message"] = (
                    f"Because the {pet_id}'s credits exceed their support obligation, "
                    f"the court should order the {resp_id} to pay the amount on line 8."
                )

        # Otherwise, the parent with less overnights pays
        #   unless their credit exceeds the amount due
        elif pet["Overnights"] > resp["Overnights"]:
            pet["RecSupport"] = 0
            resp["PTC"] = self._get_ptc(resp, data)
            resp["RecSupport"] = self._get_rec_support(resp, data["TotalSupport"])
            # Credit exceeds amount due
            if resp["RecSupport"] < 0.0:
                pet["RecSupport"] = -resp["RecSupport"]
                resp["RecSupport"] = 0.0
                data["Message"] = (
                    f"Because the {resp_id}'s credits exceed their support obligation, "
                    f"the court should order the {pet_id} to pay the amount on line 8."
                )
        elif resp["Overnights"] > pet["Overnights"]:
            resp["RecSupport"] = 0
            pet["PTC"] = self._get_ptc(pet, data)
            pet["RecSupport"] = self._get_rec_support(pet, data["TotalSupport"])
            # Credit exceeds amount due
            if pet["RecSupport"] < 0.0:
                resp["RecSupport"] = -pet["RecSupport"]
                pet["RecSupport"] = 0.0
                data["Message"] = (
                    f"Because the {pet_id}'s credits exceed their support obligation, "
                    f"the court should order the {resp_id} to pay the amount on line 8."
                )
        # If the parties equally share controlled expenses (due to equal parenting time)
        #   calculate support twice, once with each party paying controlled expenses
        #   The party which owes the larger amount pays the difference
        else:
            # Calculate as if Pet was paying controlled expenses
            resp["PTC"] = self._get_ptc(resp, data)
            resp["RecSupport"] = self._get_rec_support(resp, data["TotalSupport"])
            # Calculate as if Resp was paying controlled expenses
            pet["PTC"] = self._get_ptc(pet, data)
            pet["RecSupport"] = self._get_rec_support(pet, data["TotalSupport"])

            support_diff = resp["RecSupport"] - pet["RecSupport"]
            if support_diff > 0:
                # Resp owe more, therefore they should pay the difference
                resp["RecSupport"] = support_diff
                pet["RecSupport"] = 0
                data["Message"] = (
                    "Because the parties will equally share controlled expenses, "
                    f"the court should order the {resp_id} to pay the amount on line 8."
                )
            elif support_diff < 0:
                # Pet owes more, therefore they should pay the difference
                pet["RecSupport"] = -support_diff
                resp["RecSupport"] = 0
                data["Message"] = (
                    "Because the parties will equally share controlled expenses, "
                    f"the court should order the {pet_id} to pay the amount on line 8."
                )
            else:
                # Both parties owes the exact same amount, no one pays
                pet["RecSupport"] = 0
                resp["RecSupport"] = 0
                data["Message"] = (
                    "Because the parties will equally divide controlled expenses, "
                    "and the obligations of the parties are equal, "
                    "the court should order that neither party pays support."
                )

        data["Pet"] = pet
        data["Resp"] = resp

        return data

    def _get_basic_info(self, party, context):
        sub_child_multipliers = (
            0,
            0.065,
            0.097,
            0.122,
            0.137,
            0.146,
            0.155,
            0.164,
            0.173,
        )

        total_income = (
            (context[f"{party}GrossIncome"] or 0)
            + (context[f"{party}GrossSelfIncome"] or 0)
            + (context[f"{party}GrossNonEmp"] or 0)
        )

        result = {
            # Income from the form is MONTHLY, so it must be converted to WEEKLY
            "WeeklyIncome": _to_weekly(total_income),
            "SCCM": sub_child_multipliers[
                min(int(context[f"{party}SubChildren"] or 0), 8)
            ],
            # Child care from the form is MONTHLY, so it must be converted to WEEKLY
            "ChildCare": _to_weekly(context[f"{party}ChildCareExpense"] or 0.0),
            # No PTC by default, will be calculated later, if applicable
            "PTC": 0,
        }

        party_val = "Me" if party == "Pet" else "Spouse"
        if (
            context["ParentingScheduleYesNo"] == "Yes"
            or context["ParentingSchedule"] == "Equally divide time between parties."
        ):
            result["Overnights"] = 181
            result["IsControlledExpensePayer"] = (
                context["ControlledExpensesPayer"] == party_val
            )
        else:
            result["Overnights"] = int(context[f"{party}Overnights"])

            if (181 <= int(context["PetOvernights"]) <= 184) and (
                181 <= int(context["RespOvernights"]) <= 184
            ):
                result["IsControlledExpensePayer"] = (
                    context["ControlledExpensesPayer"] == party_val
                )
            else:
                result["IsControlledExpensePayer"] = False

        result["PtTotal"], result["PtDuplicate"] = self._get_table_pt_info(
            result["Overnights"], context["FilingState"]
        )

        if (context["MaintenancePayer"] == "Me" and party == "Pet") or (
            context["MaintenancePayer"] == "Spouse" and party == "Resp"
        ):
            # Maintenance from the form is MONTHLY, so it must be converted to WEEKLY
            result["Maintenance"] = _to_weekly(context["MaintenanceSum"])
        else:
            result["Maintenance"] = 0.0

        if (context["InsurancePayer"] == "Me" and party == "Pet") or (
            context["InsurancePayer"] == "Spouse" and party == "Resp"
        ):
            # Insurance from the form is MONTHLY, so it must be converted to WEEKLY
            result["Insurance"] = _to_weekly(context["ChildInsuranceCost"] or 0.0)
        else:
            result["Insurance"] = 0.0

        if context[f"{party}SupportCourtOrderedYesNo"] == "Yes":
            result["CSCourtOrder"] = _to_weekly(
                context[f"{party}PriorSupportAmount"] or 0.0
            )
            result["CSLegalDuty"] = 0.0
        else:
            result["CSCourtOrder"] = 0.0
            result["CSLegalDuty"] = _to_weekly(
                context[f"{party}PriorSupportAmount"] or 0.0
            )

        result["WAI"] = self._calc_wai(result)

        return result

    def _get_ptc(self, party, shared):
        percentage_wai = party["WAI"] / shared["TotalWAI"]
        avg_weekly_total_exp = shared["Obligation"] * party["PtTotal"]
        avg_weekly_dup_exp = shared["Obligation"] * party["PtDuplicate"]
        share_of_dup_exp = percentage_wai * avg_weekly_dup_exp

        return avg_weekly_total_exp - share_of_dup_exp

    def _get_rec_support(self, party, total_support):
        return (
            total_support * party["PercentageWAI"]
            - party["ChildCare"]
            - party["Insurance"]
            - party["PTC"]
        )

    def _calc_wai(self, info):
        return (
            info["WeeklyIncome"]
            - (info["WeeklyIncome"] * info["SCCM"])
            - info["CSCourtOrder"]
            - info["CSLegalDuty"]
            - info["Maintenance"]
        )

    def _calc_obligation(self, total_wai: float, state: str, num_children: int):
        schedule_file = self.data_path / "schedule.xlsx"

        if not schedule_file.exists():
            raise FileNotFoundError(f'File "{schedule_file}" does not exits!"')

        wb = load_workbook(filename=schedule_file, read_only=True)

        if state not in wb.sheetnames:
            raise Exception(f"No schedule data for state '{state}'")

        ws = wb[state]

        STEP = 10
        WAI_OFFSET = 100
        ROW_OFFSET = 3
        COL_OFFSET = 3

        # Round total_wai to the *nearest* STEP
        rounded_wai = STEP * round(total_wai / STEP)

        row = max(
            int(((rounded_wai - WAI_OFFSET) / STEP) + ROW_OFFSET),
            ROW_OFFSET,
        )

        col = int(COL_OFFSET + min(num_children, ws.max_column) - 1)

        if row >= ws.max_row:
            # Calculate obligation if total_wai is outside of table
            obligation = ws.cell(ws.max_row, col).value * total_wai
        else:
            obligation = ws.cell(row, col).value

        wb.close()

        return obligation

    def _get_table_pt_info(self, overnights, state):
        table_file = self.data_path / "table_pt.xlsx"

        if not table_file.exists():
            raise FileNotFoundError(f'File "{table_file}" does not exits!"')

        wb = load_workbook(filename=table_file, read_only=True)

        if state not in wb.sheetnames:
            raise Exception(f"No schedule data for state '{state}'")

        ws = wb[state]

        total, duplicate = 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] <= overnights:
                _, total, duplicate = row
            else:
                break

        wb.close()

        return total, duplicate


def _to_weekly(amount):
    try:
        return amount / 4
    except ZeroDivisionError:
        return 0
